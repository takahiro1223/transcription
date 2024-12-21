import os
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

## Excelのセットアップ ★１
def setupSheet(sheet):
    # セルの位置を調整
    sheet.column_dimensions['A'].width = 4
    sheet.column_dimensions['B'].width = 100
    sheet.column_dimensions['C'].width = 17
    sheet.row_dimensions[1].height = 30
    
    # セルの色・フォントを定義
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # テキスト入力
    header_cells = ['A1', 'B1', 'C1']
    header_values = ["No", "文字起こし", "音声ファイル"]
    
    # 各セルの罫線付け
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell, value in zip(header_cells, header_values):
        sheet[cell].value = value
        sheet[cell].font = header_font
        sheet[cell].fill = header_fill
        sheet[cell].alignment = header_alignment
        sheet[cell].border = thin_border

## 作業環境の用意
def create_workspace(transcription_folder_path, audio_folder_path, result_excel_path):
    # フォルダの作成
    os.makedirs(transcription_folder_path, exist_ok=True)
    os.makedirs(audio_folder_path, exist_ok=True)
    
    # Excelを作成
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "文字起こし結果"
    # Excelのセットアップ ★１
    setupSheet(sheet)
    # Excelを保存
    workbook.save(result_excel_path)

## 文字起こし結果をExcelに入力
def export_transcription_excel(transcription_results, result_excel_path, audio_folder_path):
    # 既存のExcelファイルを開く
    wb = load_workbook(result_excel_path)
    # アクティブなシートを取得
    ws = wb.active
    # 入力を始めるセルの位置
    row_counter = 2
    
    # Excelに情報を入力
    for i, segments in enumerate(transcription_results, start = 1):
        
        # Excelの1列目ににNo追加
        ws.cell(row=row_counter, column=1, value=i)
        
        # 文字起こし結果をExcelに記載
        segment_texts = "\n".join(segments)  # リストを改行で結合して文字列に変換
        ws.cell(row=row_counter, column=2, value=segment_texts)
        
        # 音声ファイルまでのパスを作成
        split_audio_files_path = os.path.join(audio_folder_path, f"音声ファイル{i}.mp3")
        # Excelの3列目に音声ファイルのリンク追加
        ws.cell(row=row_counter, column=3, value=f'=HYPERLINK("{split_audio_files_path}", "音声ファイル{i}")')
        
        # 数を数える
        row_counter += 1
        
    # Excelファイルを保存
    wb.save(result_excel_path)

## Excelにスタイルを設定
def apply_excel_styles(result_excel_path):
    wb = load_workbook(result_excel_path)
    ws = wb.active
    last_row = ws.max_row
    
    # フォントサイズ
    content_font = Font(size=10)
    # 各行のフォント位置設定
    No_alignment = Alignment(horizontal="center", vertical="center")
    transcription_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # 各セルの罫線付け
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # それぞれのセルに設定を反映
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=3):
        for cell in row:
            cell.font = content_font
            if cell.column == 1:
                cell.alignment = No_alignment
            elif cell.column == 2:
                cell.alignment = transcription_alignment
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[row[0].row].height = 50
    
    # Excelを保存
    wb.save(result_excel_path)
